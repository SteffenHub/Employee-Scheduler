import re
from typing import Dict, List
import math

import openpyxl
from openpyxl.styles import PatternFill, Side, Border, Font
from openpyxl.workbook import Workbook

from model.Team import Team
from model.Week import Week


def write_to_excel(model_result: Dict[str, bool], teams: List[Team], weeks: List[Week], shift_names: list[str]):
    workbook = Workbook()
    sheet = workbook.active
    if len(shift_names) > 3:
        print(model_result)
        [print("length of shift_names is greater than 3") for _ in range(20)]

    colors = {shift_names[0]: PatternFill(start_color='92d050', end_color='92d050', fill_type='solid'),
              shift_names[1]: PatternFill(start_color='ffc000', end_color='ffc000', fill_type='solid'),
              shift_names[2]: PatternFill(start_color='00b0f0', end_color='00b0f0', fill_type='solid'),
              "MO:M1": PatternFill(start_color='ff99ff', end_color='ff99ff', fill_type='solid'),
              "H1:M1": PatternFill(start_color='ff99ff', end_color='ff99ff', fill_type='solid'),
              "H2:M1": PatternFill(start_color='ff99ff', end_color='ff99ff', fill_type='solid'),
              "H:M2": PatternFill(start_color='99ff99', end_color='99ff99', fill_type='solid'),
              "MO:M3": PatternFill(start_color='66ffff', end_color='66ffff', fill_type='solid'),
              "H:M3": PatternFill(start_color='66ffff', end_color='66ffff', fill_type='solid'),
              "MO:M4": PatternFill(start_color='cc9900', end_color='cc9900', fill_type='solid'),
              "weekend": PatternFill(start_color='f3af9a', end_color='f3af9a', fill_type='solid'),
              "Team1": PatternFill(start_color='fff2cc', end_color='fff2cc', fill_type='solid'),
              "Team2": PatternFill(start_color='e2f0d9', end_color='e2f0d9', fill_type='solid'),
              "Team3": PatternFill(start_color='deebf7', end_color='deebf7', fill_type='solid')}
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    columns = [chr(i + 65) if i < 26 else chr(i // 26 + 64) + chr(i % 26 + 65) for i in
               range(len([day for week in weeks for day in week.days]) + 10)]
    # add Employee Names
    row = 1
    fill_to = len([day for week in weeks for day in week.days]) + 3
    for team in teams:
        for employee in team.employees:
            sheet[f"A{row * 2}"] = str(team)
            sheet[f"B{row * 2}"] = employee.name
            sheet[f"C{row * 2}"] = str([str(skill) for skill in employee.skills]).replace("[", "").replace("]",
                                                                                                           "").replace(
                "'", "")
            if employee.is_shift_manager:
                sheet[f"A{row * 2}"].font = Font(bold=True)
                sheet[f"B{row * 2}"].font = Font(bold=True)
                sheet[f"C{row * 2}"].font = Font(bold=True)
            for i in range(0, fill_to):
                sheet[f"{columns[i]}{row * 2}"].fill = colors[str(team)]
            row = row + 1

    # add name and skills
    sheet["A1"] = "Team"
    sheet["B1"] = "Name"
    sheet["C1"] = "Skills"

    # add calendar
    wochentage = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]
    column = 3
    for week in weeks:
        for j in range(1, len(week.days) + 1):
            index = (j - 1)
            sheet[f"{columns[column]}{1}"] = wochentage[index]
            if week.days[index].name == "Sa" or week.days[index].name == "Su":
                for i in range(1, len([employee for team in teams for employee in team.employees]) * 2 + 1 + 1):
                    sheet[f"{columns[column]}{i}"].fill = colors["weekend"]
                    sheet[f"{columns[column]}{i}"].border = thin_border
            column = column + 1

    # add results
    for key in model_result.keys():
        if model_result[key]:
            split = key.split("_")
            week, day, shift, team, employee, needed_skill = split[0], split[1], split[2], split[3], split[4], split[5]
            this_column = columns[
                ((wochentage.index(day) + 1) + 7 * (int(re.search(r'week(\d+)', week, re.I).group(1)) - 1) + 2)]
            this_row = ([tmp_employee.name for team in teams for tmp_employee in team.employees].index(
                employee) + 1) * 2
            sheet[
                f"{this_column}{this_row}"
            ] = shift
            sheet[
                f"{this_column}{this_row}"
            ].fill = colors[shift]
            sheet[
                f"{this_column}{this_row}"
            ].border = thin_border
            this_row = this_row + 1
            sheet[
                f"{this_column}{this_row}"
            ] = needed_skill
            sheet[
                f"{this_column}{this_row}"
            ].fill = colors[needed_skill]
            sheet[
                f"{this_column}{this_row}"
            ].border = thin_border
    workbook.save(filename="hello_world.xlsx")


def read_from_excel(name_of_excel_file: str):
    # read
    workbook = openpyxl.load_workbook(name_of_excel_file)
    sheet = workbook.active

    # get content from excel
    content = []
    for row in sheet.iter_rows(values_only=True):
        content.append(row)

    result: list[str] = []
    days = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]
    # iterate through all rows beginning with 1. first row is header containing day names
    i = 1
    while i < len(content):
        # get team und employee name
        team, name = content[i][:2]
        # start with column 3. 0-2 containing team, name and skills
        day_number = 0
        for shift, skill in zip(content[i][3:], content[i + 1][3:]):
            if shift is not None and skill is not None:
                key = f"Week{math.ceil((day_number+1)/7)}_{days[day_number%7]}_{shift}_{team}_{name}_{skill}"
                result.append(key)
                print(key)
            day_number += 1
        i += 2
    return result
